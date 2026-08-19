#!/usr/bin/env python3
"""주간 WBS 초안 생성 — 매주 월요일 KST 14:00 (A-185 · 조정자 지시 2026-08-19).

동작: 직전 WBS(reports/WBS_*.xlsm 최신)를 복사해 VBA·XLGantt 템플릿을 보존한 채
      ①해당 주 실적을 MEMORY 아카이브·git log에서 추출해 요약 ②Schedule 날짜 기준
      진척 검토 항목을 표시 ③reports/WBS_{MMDD}.xlsm 저장.

⚠️ 작업 행 구성·진척률은 **사람 검토 대상**이다. 이 스크립트는 초안 파일과 '검토 필요'
   목록을 만들 뿐, 실적 수치를 자동으로 확정하지 않는다(과장 방지 — CLAUDE.md §6 정신).

출력: reports/WBS_{MMDD}.xlsm + stdout 요약(워크플로우가 STEP_SUMMARY로 노출)
"""
from __future__ import annotations

import glob
import os
import re
import shutil
import subprocess
from datetime import date, timedelta
from pathlib import Path

import openpyxl

REPORTS = Path("reports")
MEMORY = Path("docs/memory_archive")


def _latest_wbs() -> Path:
    files = sorted(glob.glob("reports/WBS_*.xlsm"))
    if not files:
        raise SystemExit("[오류] 기준 WBS 파일 없음 — reports/WBS_*.xlsm 필요")
    return Path(files[-1])


def _week_range(today: date) -> tuple[date, date]:
    """보고 주간 = 직전 목요일 ~ 이번 수요일 (월요일 14시 보고 기준 직전 7일)."""
    return today - timedelta(days=7), today - timedelta(days=1)


def _memory_entries(since: date, until: date) -> list[str]:
    """월별 아카이브에서 기간 내 행(| YYYY-MM-DD | ID | …)을 뽑는다."""
    out: list[str] = []
    for f in sorted(MEMORY.glob("*.md")):
        for line in f.read_text(encoding="utf-8").splitlines():
            m = re.match(r"\|\s*(\d{4}-\d{2}-\d{2})\s*\|\s*([A-Z]+-\d+)\s*\|\s*([^|]+)\|\s*([^|]+)\|",
                         line)
            if not m:
                continue
            try:
                d = date.fromisoformat(m.group(1))
            except ValueError:
                continue
            if since <= d <= until:
                out.append(f"{m.group(1)} [{m.group(2)}] {m.group(3).strip()} — {m.group(4).strip()[:90]}")
    return out


def _git_log(since: date, until: date) -> list[str]:
    try:
        r = subprocess.run(
            ["git", "log", f"--since={since}", f"--until={until + timedelta(days=1)}",
             "--date=short", "--pretty=format:%ad %h %s"],
            capture_output=True, text=True, timeout=60)
        return [l for l in r.stdout.splitlines() if l.strip()]
    except Exception as e:
        print(f"[경고] git log 실패: {e}")
        return []


def main() -> int:
    today = date.today()
    since, until = _week_range(today)
    src = _latest_wbs()
    dst = REPORTS / f"WBS_{today:%m%d}.xlsm"

    if dst.exists():
        print(f"[정보] {dst.name} 이미 존재 — 덮어쓰지 않고 종료(중복 실행 방지)")
        return 0

    shutil.copyfile(src, dst)
    wb = openpyxl.load_workbook(dst, keep_vba=True, data_only=False)
    ws = wb["Schedule"]

    # 기한 경과·임박 작업을 검토 대상으로 표시
    review: list[str] = []
    for r in range(5, 57):
        wbs_no, q, ae = ws.cell(r, 4).value, ws.cell(r, 17).value, ws.cell(r, 31).value
        name = next((ws.cell(r, c).value for c in (5, 6, 7) if ws.cell(r, c).value), None)
        if not wbs_no or not q:
            continue
        qd = q.date() if hasattr(q, "date") else q
        pct = float(ae) if isinstance(ae, (int, float)) else 0.0
        if qd < today and pct < 1.0:
            review.append(f"🔴 {wbs_no} {name} — 기한 {qd} 경과, 진척 {pct:.0%}")
        elif qd <= today + timedelta(days=7) and pct < 1.0:
            review.append(f"🟡 {wbs_no} {name} — 기한 {qd} 임박, 진척 {pct:.0%}")

    wb.save(dst)

    mem = _memory_entries(since, until)
    commits = _git_log(since, until)

    print(f"# 주간 WBS 초안 — {dst.name}\n")
    print(f"- 기준 파일: `{src.name}` → 복사 생성 (VBA·XLGantt 보존)")
    print(f"- 보고 주간: {since} ~ {until}")
    print(f"- MEMORY 기록: **{len(mem)}건** · 커밋: **{len(commits)}건**\n")
    print("## 검토 필요 (기한 경과·임박)\n")
    print("\n".join(f"- {x}" for x in review) if review else "- 없음")
    print(f"\n## 금주 MEMORY 기록 ({len(mem)}건)\n")
    print("\n".join(f"- {x}" for x in mem[:40]) if mem else "- 없음")
    if len(mem) > 40:
        print(f"- … 외 {len(mem)-40}건")
    print(f"\n## 커밋 ({len(commits)}건)\n")
    print("\n".join(f"- {c}" for c in commits[:30]) if commits else "- 없음")
    print("\n> ⚠️ 작업 행 구성·진척률은 **자동 확정하지 않음** — C-01 검토 후 갱신 필요.")
    print(f"\n[완료] {dst}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
